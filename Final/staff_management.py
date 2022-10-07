import tkinter as tk
import re
from tkinter import ttk
from tkinter import messagebox as msb
from openpyxl import load_workbook

def check_blank_data():
    errors = []
    message=''
    if entry_value_id.get()=='':
        errors.append('Id')
    if entry_value_age.get()=='':
        errors.append('Age')
    if entry_value_full_name.get()=='':
        errors.append('Full_name')
    if entry_value_email.get()=='':
        errors.append('Email')
    if entry_value_phone.get()=='':
        errors.append('Phone')
    if cbx_selected_department.get()=='--- Lựa chọn phòng ban ---':
        errors.append('Department')
    for error_index in range(0,len(errors)):
        if error_index == len(errors)-1:
            message += errors[error_index] + ' '
        else:    
            message += errors[error_index] + ', '
    if message =='':
        return True
    else:
        msb.showwarning('Warnning',message+'is blank!')
        return False  

def validate_phone(value):
    # Bắt lỗi nhập số điện thoại
    # '^0': ký tự đầu phải là 0; 'd{9}': yêu cầu phải là số 
    pattern = '^0\d{9}'
    if entry_value_phone.get() !='':
        if re.fullmatch(pattern, entry_value_phone.get()) is None:
            return False
        else:
            return True
    
def validate_age(value):
    pattern = '\d+'
    if entry_value_age.get() !='':
        if re.fullmatch(pattern, entry_value_age.get()) is None:
            return False
        else:
            return True


def save_data():
    wb.save('staff_list.xlsx')

def clear_data():
    entry_value_id.set('')
    entry_value_age.set('')
    entry_value_full_name.set('')
    entry_value_email.set('')
    entry_value_phone.set('')
    cbx_department.current(0)

def item_selected():
    selected_item = tvw_info.selection()
    item = tvw_info.item(selected_item,option='values')
    return item

def view_staff():
    try:
        item = item_selected()
        entry_value_id.set(item[0])
        entry_value_full_name.set(item[1])
        entry_value_email.set(item[2])
        entry_value_age.set(item[3])
        entry_value_phone.set(item[4])
        cbx_department.current(cbx_department['value'].index(item[5]))
    except:
        clear_data()

def view_all_staff(): 
    global wb
    global ws
    
    wb = load_workbook('staff_list.xlsx')
    ws = wb['Sheet1']
    
    for item in tvw_info.get_children():
        tvw_info.delete(item)
    for row in range(2,ws.max_row+1):
        tvw_value = []
        for col in range(1,ws.max_column+1):
            tvw_value.append(ws.cell(row=row,column=col).value)
        tvw_info.insert('',tk.END,value=tvw_value)

def add_staff():
    new_id=entry_value_id.get()    
    new_age = ''
    new_age=entry_value_age.get()    
    new_full_name=entry_value_full_name.get()    
    new_email=entry_value_email.get()    
    new_phone = ''
    new_phone=entry_value_phone.get()    
    new_deparment=cbx_selected_department.get()    
    new_staff=[new_id,new_full_name,new_email,new_age,new_phone,new_deparment]

    if check_blank_data() == True:        
        # Kiểm tra xem ID có bị trùng không
        for row in ws:
            if row[0].row > 1:
                if row[0].value == new_id:
                    print(row[0].value)
                    msb.showwarning('Warnning','MÃ ĐÃ TỒN TẠI!')
                    return 
        # Thực hiện thêm vào file excel nếu không trùng
        last_row = ws.max_row
        last_column = 6
        for col in range(1,last_column+1):
            ws.cell(row=last_row+1,column=col).value=new_staff[col-1]
        msb.showinfo('Anoucement','THÊM THÀNH CÔNG!')
        save_data()
        clear_data()
        view_all_staff()
    
def update_staff():
    global wb
    global ws 
    
    update_id = entry_value_id.get()
    update_age = ''
    update_age=entry_value_age.get()    
    update_full_name=entry_value_full_name.get()    
    update_email=entry_value_email.get()    
    update_phone = ''
    update_phone=entry_value_phone.get()    
    update_department=cbx_selected_department.get()    
    update_staff=[update_id,update_full_name,update_email,update_age,update_phone,update_department]
    
    for row in ws:
        if row[0].value == update_id:
            for col in range(0,ws.max_column):
                row[col].value=update_staff[col]
                
    msb.showinfo('Anoucement','CẬP NHẬT THÀNH CÔNG!')
    save_data()
    clear_data()        
    wb = load_workbook('staff_list.xlsx')
    ws = wb['Sheet1']
    view_all_staff()
    
def delete_staff():
    item = item_selected()
    answer = msb.askyesno('Confirm','BẠN CÓ MUỐN XÓA KHÔNG?')
    if answer == 1:
        for row in ws:
            if row[0].value == item[0]:
                ws.delete_rows(row[0].row) 
                save_data()
                msb.showinfo('Anoucement','XÓA THÀNH CÔNG!')        
    view_all_staff()

def exit():
    window.destroy()

# Khởi tạo window chính
window = tk.Tk()
window.title('STAFF MANAGEMENT')
window.geometry('950x450')
window.resizable(False,False)

# Khởi tạo đối tượng đọc file excel
wb = load_workbook('staff_list.xlsx')
ws = wb['Sheet1']

# Khởi tạo biến variable của các widget
entry_value_id = tk.StringVar()
entry_value_full_name = tk.StringVar()
entry_value_email = tk.StringVar()
entry_value_age = tk.StringVar()
entry_value_phone = tk.StringVar()
cbx_selected_department = tk.StringVar()
info_column = ('id','full_name','email','age','phone','department')

# Khởi tạo frame
frame1 = ttk.Frame(window)


# Khởi tạo các trường để nhập thông tin
frame_input_info = ttk.Labelframe(frame1,text='Nhập thông tin')

label_id = ttk.Label(frame_input_info,text='ID')
label_id.grid(row=0,column=0,sticky='e',padx='5',pady='5')

label_age = ttk.Label(frame_input_info,text='Age')
label_age.grid(row=0,column=2,sticky='e',padx='5',pady='5')

label_full_name = ttk.Label(frame_input_info,text='Full Name')
label_full_name.grid(row=1,column=0,sticky='e',padx='5',pady='5')

label_email = ttk.Label(frame_input_info,text='Email')
label_email.grid(row=1,column=2,sticky='e',padx='5',pady='5')

label_phone = ttk.Label(frame_input_info,text='Phone.No')
label_phone.grid(row=2,column=0,sticky='e',padx='5',pady='5')

label_Department = ttk.Label(frame_input_info,text='Department')
label_Department.grid(row=2,column=2,sticky='e',padx='5',pady='5')

entry_id = ttk.Entry(frame_input_info,textvariable=entry_value_id)
entry_id.grid(row=0,column=1,sticky='w',pady='5')

entry_age = ttk.Entry(frame_input_info,textvariable=entry_value_age)
entry_age.grid(row=0,column=3,sticky='w',pady='5')


entry_full_name = ttk.Entry(frame_input_info,textvariable=entry_value_full_name)
entry_full_name.grid(row=1,column=1,sticky='w',pady='5')

entry_email = ttk.Entry(frame_input_info,textvariable=entry_value_email)
entry_email.grid(row=1,column=3,sticky='w',pady='5')

entry_phone = ttk.Entry(frame_input_info,textvariable=entry_value_phone)
entry_phone.grid(row=2,column=1,sticky='w',pady='5')

cbx_department = ttk.Combobox(frame_input_info,state= "readonly",textvariable=cbx_selected_department)
cbx_department['value']=('--- Select a Department ---','Kế toán','Giám đốc','Bán hàng')
cbx_department.current(0)
cbx_department.grid(row=2,column=3,sticky='w',pady='5')

# Khởi tạo các Button để thực hiện các chức năng
frame_function = ttk.Labelframe(frame1,text='Chức năng')

btn_view = ttk.Button(frame_function, text='View All', command=view_all_staff)
btn_view.grid(row=0,column=0,sticky='w',padx='5',pady='5')

btn_add = ttk.Button(frame_function, text='Add', command=add_staff)
btn_add.grid(row=0,column=1,sticky='w',padx='5',pady='5')

btn_update = ttk.Button(frame_function, text='Update', command=update_staff)
btn_update.grid(row=0,column=2,sticky='w',padx='5',pady='5')

btn_delete = ttk.Button(frame_function, text='Delete', command=delete_staff)
btn_delete.grid(row=0,column=3,sticky='w',padx='5',pady='5')

btn_get_info = ttk.Button(frame_function, text='Get Info', command=view_staff)
btn_get_info.grid(row=1,column=0,sticky='w',padx='5',pady='5')

btn_clear = ttk.Button(frame_function, text='Clear', command=clear_data)
btn_clear.grid(row=1,column=1,sticky='w',padx='5',pady='5')

btn_exit = ttk.Button(frame_function, text='Exit', command=exit)
btn_exit.grid(row=0,column=4,sticky='w',padx='5',pady='5')

# Frame hiển thị thông tin nhân viên
frame_display_info = ttk.Labelframe(window, text='Thông tin nhân viên')

tvw_info = ttk.Treeview(frame_display_info,column=info_column,show='headings')
tvw_info.column('id',width=30)
tvw_info.column('full_name',width=100)
tvw_info.column('email',width=100)
tvw_info.column('age',width=30)
tvw_info.column('phone',width=80)
tvw_info.column('department',width=50)

tvw_info.heading('id',text='ID')
tvw_info.heading('full_name',text='Full Name')
tvw_info.heading('email',text='Email')
tvw_info.heading('age',text='Age')
tvw_info.heading('phone',text='Phone.No')
tvw_info.heading('department',text='Department')

tvw_info.grid(row=1,column=0,sticky='w',padx='5',pady='5')
tvw_info.bind('<<TreeviewSelect>>',item_selected())


# Sắp xếp các frame
frame_input_info.pack(padx='10',pady='20', anchor='w')
frame_function.pack(padx='10',pady='20', anchor='w')
frame1.pack(padx='10',pady='20', anchor='w',side='left')
frame_display_info.pack(padx='10',pady='100', anchor='w',side='right')



window.mainloop()